import json
import openpyxl
import random
from environment import FleetAIEnv
from environment.models import ResetRequest, InspectorAction
from environment.error_injector import ErrorInjector
from environment.tickets import TicketRecord, TicketGroundTruth
from environment.graders import grade_inspector
from environment.models import WorkerResponse, InjectedError, GraderResult, InspectorObservation

DATASET_PATH = r"C:\Users\kavan\Downloads\amazon_error_filled_tickets_dataset.xlsx"

ISSUE_TYPE_TO_CLASSIFICATION = {
    "Delivery": "shipping", "Payment": "billing", "Login": "account",
    "Product": "product", "Refund": "billing", "Account": "account",
    "Shipping": "shipping", "Return": "product", "Technical": "technical",
    "Prime": "account",
}

PRIORITY_MAP = {
    "High": "high", "high": "high", "Medium": "medium", "medium": "medium",
    "Low": "low", "low": "low", "Urgent": "urgent", "urgent": "urgent",
    "Critical": "urgent",
}

TIER_MAP = {"prime": "gold", "non-prime": "silver", "new": "bronze"}


def classify_sentiment(description, notes):
    text = f"{description or ''} {notes or ''}".lower()
    negative_words = ["angry", "frustrated", "terrible", "worst", "refund", "complaint",
                      "overcharged", "not received", "never", "stolen", "broken", "damage"]
    if any(w in text for w in negative_words):
        return "negative"
    return "neutral"


def build_ticket_and_worker(row_data, headers, rng):
    d = dict(zip(headers, row_data))
    classification = ISSUE_TYPE_TO_CLASSIFICATION.get(d.get("issue_type", ""), "general")
    priority_raw = str(d.get("priority", "")).strip()
    priority = PRIORITY_MAP.get(priority_raw, "medium")

    ticket_id = str(d.get("ticket_id") or "UNKNOWN")
    subject = str(d.get("sub_issue") or d.get("description") or "")[:80]
    body = str(d.get("description") or "")
    customer_id = str(d.get("customer_id") or "C-000")
    issue_type = str(d.get("issue_type") or "General")

    gt = TicketGroundTruth(
        category=classification,
        priority=priority,
        department="general_support",
        difficulty=1,
        key_terms=[],
        required_response_elements=[],
        expected_resolution_actions=[],
        escalation_required=False,
        sentiment_should_address=False,
    )

    ticket = TicketRecord(
        id=ticket_id,
        subject=f"[{issue_type}] {subject}",
        body=body,
        customer_id=customer_id,
        customer_name=customer_id,
        customer_tier=TIER_MAP.get(str(d.get("notes", "")).lower(), "silver"),
        customer_tenure_days=180,
        sentiment=classify_sentiment(d.get("description"), d.get("notes")),
        previous_ticket_count=2,
        ground_truth=gt,
    )

    injector = ErrorInjector(rng=rng)
    worker, errors = injector.generate_worker_response(ticket)

    return ticket, worker, errors, d.get("notes", "")


def build_heuristic_inspection(ticket, worker):
    flagged = False
    flagged_fields = []
    issues = []
    suggested = {}

    sentiment = (ticket.sentiment or "neutral").lower()
    w_resp = (worker.response or "").lower()
    w_priority = (worker.priority or "").lower()

    if sentiment in ("negative", "very_negative"):
        has_apology = any(w in w_resp for w in ["sorry", "apologize", "apologies", "regret", "inconvenience"])
        if not has_apology:
            flagged = True
            flagged_fields.append("response")
            issues.append({"field": "response", "reason": f"Customer sentiment is {sentiment} but no apology in response"})

    body_lower = f"{(ticket.subject or '').lower()} {(ticket.body or '').lower()}"
    urgent_keywords = ["urgent", "immediately", "asap", "emergency", "critical"]
    has_urgency = any(kw in body_lower for kw in urgent_keywords)
    if has_urgency and w_priority in ("low", "medium"):
        flagged = True
        flagged_fields.append("priority")
        issues.append({"field": "priority", "reason": "Ticket contains urgency language but priority is too low"})
        suggested["priority"] = "urgent"

    if not worker.resolution_actions or len(worker.resolution_actions) == 0:
        flagged = True
        flagged_fields.append("resolution_actions")
        issues.append({"field": "resolution_actions", "reason": "No resolution actions provided"})

    return InspectorAction(
        flagged=flagged,
        flagged_fields=flagged_fields,
        issues=issues,
        suggested_corrections=suggested,
        confidence=0.6 if flagged else 0.2,
    )


def grade_ticket(ticket, worker, action, injected_errors):
    observation = InspectorObservation(
        ticket=ticket,
        worker_response=WorkerResponse(
            classification=worker.classification,
            priority=worker.priority,
            department=worker.department,
            response=worker.response,
            resolution_actions=worker.resolution_actions,
        ),
        instructions="Review the worker response for errors.",
        hints=[],
        step=1,
        max_steps=3,
    )

    grader_result = grade_inspector(
        ticket=ticket,
        worker=worker,
        action=action,
        injected_errors=injected_errors,
    )
    return grader_result


def main():
    print("=" * 80)
    print("  FLEETAI - AMAZON TICKET DATASET TEST")
    print("=" * 80)

    wb = openpyxl.load_workbook(DATASET_PATH)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    rows = list(ws.iter_rows(min_row=2, values_only=True))

    print(f"  Dataset: {DATASET_PATH}")
    print(f"  Tickets: {len(rows)}")
    print(f"  Columns: {len(headers)}")
    print("=" * 80)

    results = []
    total_score = 0
    passed = 0
    failed = 0

    for i, row in enumerate(rows):
        try:
            ticket_id = str(row[headers.index("ticket_id")] or "UNKNOWN")
            issue_type = str(row[headers.index("issue_type")] or "General")
            sub_issue = str(row[headers.index("sub_issue")] or "N/A")
            notes = str(row[headers.index("notes")] or "")
        except (ValueError, IndexError):
            ticket_id = f"ROW-{i+1}"
            issue_type = "Unknown"
            sub_issue = "Unknown"
            notes = ""

        rng = random.Random(i + 42)
        ticket, worker, injected_errors, ticket_notes = build_ticket_and_worker(row, headers, rng)
        action = build_heuristic_inspection(ticket, worker)

        try:
            grader_result = grade_inspector(injected_errors, action, ticket)
            score = grader_result.overall_score
        except Exception as e:
            score = 0.0
            grader_result = None

        total_score += score
        status = "PASS" if score > 0.3 else "FAIL"
        if score > 0.3:
            passed += 1
        else:
            failed += 1

        results.append({
            "ticket_id": ticket_id,
            "issue_type": issue_type,
            "sub_issue": sub_issue,
            "notes": notes,
            "flagged": action.flagged,
            "flagged_fields": action.flagged_fields,
            "injected_errors": [e.field for e in injected_errors],
            "score": round(score, 4),
            "status": status,
        })

        err_str = ", ".join(e.field for e in injected_errors) if injected_errors else "none"
        flag_icon = "FLAG" if action.flagged else "OK  "
        print(f"  [{i+1:2d}] {ticket_id:12s} | {issue_type:12s} | {sub_issue:20s} | {flag_icon} | errors={err_str:20s} | {score:.3f} | {status}")

    print()
    print("=" * 80)
    print(f"  RESULTS SUMMARY")
    print(f"{'=' * 80}")
    print(f"  Total tickets:  {len(results)}")
    print(f"  Passed:         {passed}")
    print(f"  Failed:         {failed}")
    print(f"  Pass rate:      {passed/len(results)*100:.1f}%")
    print(f"  Avg score:      {total_score/len(results):.4f}")

    issue_types = {}
    for r in results:
        it = r["issue_type"]
        if it not in issue_types:
            issue_types[it] = {"scores": [], "pass": 0, "fail": 0}
        issue_types[it]["scores"].append(r["score"])
        if r["status"] == "PASS":
            issue_types[it]["pass"] += 1
        else:
            issue_types[it]["fail"] += 1

    print()
    print(f"  {'Issue Type':<15s} {'Count':>6s} {'Pass':>6s} {'Fail':>6s} {'Avg Score':>10s}")
    print(f"  {'-'*15} {'-'*6} {'-'*6} {'-'*6} {'-'*10}")
    for it, data in sorted(issue_types.items()):
        avg = sum(data["scores"]) / len(data["scores"])
        print(f"  {it:<15s} {len(data['scores']):>6d} {data['pass']:>6d} {data['fail']:>6d} {avg:>10.4f}")

    output_path = "amazon_dataset_test_results.json"
    with open(output_path, "w") as f:
        json.dump({
            "total": len(results),
            "passed": passed,
            "failed": failed,
            "pass_rate": round(passed / len(results) * 100, 1),
            "avg_score": round(total_score / len(results), 4),
            "by_issue_type": {
                it: {"count": len(d["scores"]), "pass": d["pass"], "fail": d["fail"],
                      "avg_score": round(sum(d["scores"]) / len(d["scores"]), 4)}
                for it, d in issue_types.items()
            },
            "results": results,
        }, f, indent=2)
    print(f"\n  Results saved to {output_path}")


if __name__ == "__main__":
    main()
