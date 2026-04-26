"""
FleetAI - Scalable Oversight Dashboard

Run:  python demo_ui.py
Open: http://localhost:7862
"""

import json
import random
import time
from typing import Any, Dict, List, Optional, Tuple

import gradio as gr

from environment import FleetAIEnv
from environment.models import ResetRequest, InspectorAction

TASK_LABELS = {
    "inspection_easy": "Easy (Obvious Errors)",
    "inspection_hard": "Hard (Subtle Errors + Clean Traps)",
    "inspection_adversarial": "Adversarial (Designed to Trick)",
}

TASK_DESCRIPTIONS = {
    "inspection_easy": "Worker responses contain obvious errors. Good for getting started.",
    "inspection_hard": "Responses may contain subtle errors or be completely clean. Requires careful analysis.",
    "inspection_adversarial": "Many responses are intentionally correct but designed to look suspicious.",
}

TASK_COLORS = {"inspection_easy": "#22c55e", "inspection_hard": "#f59e0b", "inspection_adversarial": "#ef4444"}

ISSUE_TYPE_TO_CLASSIFICATION = {
    "Delivery": "shipping", "Payment": "billing", "Login": "account",
    "Product": "product", "Refund": "billing", "Account": "account",
    "Shipping": "shipping", "Return": "product", "Technical": "technical",
    "Prime": "account", "General": "general",
}

PRIORITY_MAP = {
    "High": "high", "high": "high", "Medium": "medium", "medium": "medium",
    "Low": "low", "low": "low", "Urgent": "urgent", "urgent": "urgent",
    "Critical": "urgent",
}

TIER_MAP = {"prime": "gold", "non-prime": "silver", "new": "bronze"}

env = FleetAIEnv()
state = {
    "observation": None, "info": None, "step": 0, "done": False,
    "best_reward": 0.0, "history": [],
    "total_inspected": 0, "total_passed": 0, "total_failed": 0,
    "score_sum": 0.0, "session_start": time.time(), "notifications": [],
}


def notify(message, level="info"):
    state["notifications"].insert(0, {"time": time.strftime("%H:%M:%S"), "message": message, "level": level})
    if len(state["notifications"]) > 50:
        state["notifications"] = state["notifications"][:50]


def _build_heuristic_action(ticket, worker):
    flagged = False
    flagged_fields = []
    issues = []
    suggested = {}
    sentiment = (ticket.sentiment or "neutral").lower()
    w_resp = (worker.response or "").lower()
    w_priority = (worker.priority or "").lower()
    w_actions = worker.resolution_actions or []
    if sentiment in ("negative", "very_negative"):
        has_apology = any(w in w_resp for w in ["sorry", "apologize", "apologies", "regret", "inconvenience"])
        if not has_apology:
            flagged = True
            flagged_fields.append("response")
            issues.append({"field": "response", "reason": f"Customer sentiment is {sentiment} but no apology"})
            suggested["response"] = f"Dear {ticket.customer_name}, I sincerely apologize..."
    body_lower = f"{(ticket.subject or '').lower()} {(ticket.body or '').lower()}"
    urgent_keywords = ["urgent", "immediately", "asap", "emergency", "critical"]
    if any(kw in body_lower for kw in urgent_keywords) and w_priority in ("low", "medium"):
        flagged = True
        flagged_fields.append("priority")
        issues.append({"field": "priority", "reason": "Priority too low for urgency language"})
        suggested["priority"] = "urgent"
    customer_name_last = (ticket.customer_name or "").split()[-1].lower() if ticket.customer_name else ""
    if customer_name_last and customer_name_last not in w_resp:
        flagged = True
        flagged_fields.append("response")
        issues.append({"field": "response", "reason": f"Not personalized for '{ticket.customer_name}'"})
    if not w_actions:
        flagged = True
        flagged_fields.append("resolution_actions")
        issues.append({"field": "resolution_actions", "reason": "No resolution actions provided"})
    if not flagged:
        flagged_fields, issues, suggested = [], [], {}
    return InspectorAction(flagged=flagged, flagged_fields=flagged_fields, issues=issues, suggested_corrections=suggested, confidence=0.6 if flagged else 0.2)


def card(title, content, border_color="#3b82f6"):
    return f'<div style="padding:16px; border-radius:12px; background:#1a1a2e; border:1px solid #2a2a4a; border-left:4px solid {border_color}; margin-bottom:12px;"><h4 style="margin:0 0 10px 0; color:#e2e8f0; font-size:15px;">{title}</h4>{content}</div>'


def metric_box(value, label, color="#3b82f6"):
    return f'<div style="text-align:center; padding:12px; background:#12122a; border-radius:10px; border:1px solid #2a2a4a;"><div style="font-size:24px; font-weight:700; color:{color};">{value}</div><div style="color:#8888aa; font-size:11px; margin-top:2px;">{label}</div></div>'


def progress_bar(value, color="#3b82f6", height="8px"):
    pct = max(0, min(100, int(value * 100)))
    return f'<div style="background:#12122a; border-radius:4px; height:{height}; overflow:hidden;"><div style="background:{color}; height:100%; width:{pct}%; border-radius:4px; transition:width 0.3s;"></div></div>'


def start_episode(task, seed):
    state["done"] = False
    state["step"] = 0
    state["best_reward"] = 0.0
    result = env.reset(ResetRequest(task=task, seed=seed))
    obs = result.observation
    info = result.info
    state["observation"] = obs
    state["info"] = info

    ticket = obs.ticket
    worker = obs.worker_response
    sent_emoji = {"negative": "😡", "very_negative": "🤬", "neutral": "😐", "positive": "😊"}.get(ticket.sentiment, "😐")

    ticket_content = (
        f'<div style="display:grid; grid-template-columns:1fr 1fr; gap:8px; margin-bottom:10px;">'
        f'<div style="padding:8px; background:#12122a; border-radius:8px;"><span style="color:#8888aa; font-size:11px;">ID</span><br><span style="color:#e2e8f0; font-weight:600;">{ticket.id}</span></div>'
        f'<div style="padding:8px; background:#12122a; border-radius:8px;"><span style="color:#8888aa; font-size:11px;">Customer</span><br><span style="color:#e2e8f0; font-weight:600;">{ticket.customer_name}</span></div>'
        f'<div style="padding:8px; background:#12122a; border-radius:8px;"><span style="color:#8888aa; font-size:11px;">Tier</span><br><span style="color:#fbbf24; font-weight:600;">{ticket.customer_tier.title()}</span></div>'
        f'<div style="padding:8px; background:#12122a; border-radius:8px;"><span style="color:#8888aa; font-size:11px;">Sentiment</span><br><span style="font-size:16px;">{sent_emoji} {ticket.sentiment.title()}</span></div></div>'
        f'<div style="padding:10px; background:#12122a; border-radius:8px; margin-bottom:6px;"><span style="color:#8888aa; font-size:11px;">Description</span><br><span style="color:#e2e8f0;">{ticket.body}</span></div>'
        f'<div style="color:#8888aa; font-size:11px;">Previous tickets: {ticket.previous_ticket_count} | <span style="color:{TASK_COLORS.get(task, "#8888aa")};">{TASK_LABELS.get(task, task)}</span></div>'
    )

    fail_html = '<span style="color:#ef4444;">❌ None</span>'
    priority_color = "#ef4444" if worker.priority == "urgent" else "#f59e0b" if worker.priority == "high" else "#22c55e"
    worker_content = (
        f'<div style="display:grid; grid-template-columns:1fr 1fr 1fr; gap:8px; margin-bottom:10px;">'
        f'<div style="padding:8px; background:#12122a; border-radius:8px;"><span style="color:#8888aa; font-size:11px;">Classification</span><br><span style="color:#a78bfa; font-weight:600;">{worker.classification or "N/A"}</span></div>'
        f'<div style="padding:8px; background:#12122a; border-radius:8px;"><span style="color:#8888aa; font-size:11px;">Priority</span><br><span style="color:{priority_color}; font-weight:600;">{worker.priority or "N/A"}</span></div>'
        f'<div style="padding:8px; background:#12122a; border-radius:8px;"><span style="color:#8888aa; font-size:11px;">Department</span><br><span style="color:#38bdf8; font-weight:600;">{worker.department or "N/A"}</span></div></div>'
        f'<div style="padding:10px; background:#12122a; border-radius:8px; margin-bottom:6px;"><span style="color:#8888aa; font-size:11px;">Response</span><br><span style="color:#e2e8f0;">{(worker.response or "N/A")[:400]}</span></div>'
        f'<div style="padding:8px; background:#12122a; border-radius:8px;"><span style="color:#8888aa; font-size:11px;">Resolution Actions</span><br><span style="color:#e2e8f0;">{", ".join(worker.resolution_actions) if worker.resolution_actions else fail_html}</span></div>'
    )

    instructions = obs.instructions if obs.instructions else ""

    notify(f"Episode started: {ticket.id}", "success")
    return (
        card("Ticket", ticket_content, "#3b82f6"),
        card("Worker Response", worker_content, "#8b5cf6"),
        instructions,
        gr.update(visible=True),
        gr.update(value=""),
        gr.update(value=""),
    )


def submit_inspection(flagged, flagged_fields, issues_json, corrections_json, confidence):
    if state["observation"] is None:
        return gr.update(visible=True), gr.update(value="❌ Start an episode first!")
    if state["done"]:
        return gr.update(visible=True), gr.update(value="ℹ️ Episode complete. Start a new episode.")

    parsed_fields = [f.strip() for f in flagged_fields.split(",") if f.strip()] if flagged_fields.strip() else []
    if flagged and not parsed_fields:
        return gr.update(visible=True), gr.update(value="⚠️ Flagged=True but no fields listed. Add fields or uncheck Flagged.")

    parsed_issues = []
    if issues_json.strip():
        try:
            parsed_issues = json.loads(issues_json.strip())
        except json.JSONDecodeError as e:
            return gr.update(visible=True), gr.update(value=f"❌ Invalid issues JSON: {e}")

    parsed_corrections = {}
    if corrections_json.strip():
        try:
            parsed_corrections = json.loads(corrections_json.strip())
        except json.JSONDecodeError as e:
            return gr.update(visible=True), gr.update(value=f"❌ Invalid corrections JSON: {e}")

    action = InspectorAction(flagged=flagged, flagged_fields=parsed_fields, issues=parsed_issues, suggested_corrections=parsed_corrections, confidence=confidence)
    step_result = env.step(action)
    reward = step_result.reward
    info = step_result.info
    state["step"] += 1
    if reward > state["best_reward"]:
        state["best_reward"] = reward

    grader = info.get("grader", {})
    components = grader.get("component_scores", grader.get("components", []))

    if reward >= 0.8:
        status_label, status_color = "🏆 EXCELLENT", "#22c55e"
    elif reward >= 0.6:
        status_label, status_color = "✅ GOOD", "#22c55e"
    elif reward >= 0.4:
        status_label, status_color = "⚠️ OK", "#f59e0b"
    elif reward >= 0.2:
        status_label, status_color = "❌ NEEDS WORK", "#f59e0b"
    else:
        status_label, status_color = "💀 FAIL", "#ef4444"

    comp_html = ""
    for cs in components:
        name = cs.get("name", "") if isinstance(cs, dict) else getattr(cs, "name", "")
        score = cs.get("score", 0) if isinstance(cs, dict) else getattr(cs, "score", 0)
        bar_c = "#22c55e" if score >= 0.7 else "#f59e0b" if score >= 0.4 else "#ef4444"
        comp_html += f'<div style="margin-bottom:6px;"><div style="display:flex; justify-content:space-between; margin-bottom:2px;"><span style="color:#e2e8f0; font-size:12px;">{name.replace("_", " ").title()}</span><span style="color:#8888aa; font-size:11px;">{score:.2f}</span></div>{progress_bar(score, bar_c)}</div>'

    injected_errors = info.get("injected_errors", [])
    err_rows = ""
    for e in injected_errors:
        field = e.get("field", "") if isinstance(e, dict) else getattr(e, "field", "")
        gt_val = e.get("ground_truth", "") if isinstance(e, dict) else ""
        corrupted = e.get("corrupted", "") if isinstance(e, dict) else ""
        err_rows += f'<tr><td style="padding:4px 8px; color:#e2e8f0; border-bottom:1px solid #2a2a4a;">{field}</td><td style="padding:4px 8px; color:#22c55e; border-bottom:1px solid #2a2a4a;">{gt_val}</td><td style="padding:4px 8px; color:#ef4444; border-bottom:1px solid #2a2a4a;">{corrupted}</td></tr>'
    if not err_rows:
        err_rows = '<tr><td colspan="3" style="padding:6px; color:#22c55e; text-align:center;">✅ No errors (clean response)</td></tr>'

    results_html = (
        f'<div style="text-align:center; padding:16px; background:#12122a; border-radius:10px; margin-bottom:12px;">'
        f'<div style="font-size:32px; font-weight:700; color:{status_color};">{reward:.2f}</div>'
        f'<div style="color:#8888aa; font-size:13px;">{status_label} (Step {state["step"]}/3)</div></div>'
        f'{comp_html}'
        f'<div style="margin-top:8px;"><span style="color:#8888aa; font-size:11px;">INJECTED ERRORS</span>'
        f'<table style="width:100%; border-collapse:collapse; font-size:12px; margin-top:4px;"><tr style="color:#8888aa;"><th style="text-align:left; padding:4px 8px;">Field</th><th style="text-align:left; padding:4px 8px;">Correct</th><th style="text-align:left; padding:4px 8px;">Corrupted</th></tr>{err_rows}</table></div>'
    )

    action_md = f"```json\n{json.dumps(action.model_dump(), indent=2)}\n```"

    state["total_inspected"] += 1
    state["score_sum"] += reward
    if reward >= 0.4:
        state["total_passed"] += 1
    else:
        state["total_failed"] += 1
    state["history"].append({"ticket_id": state["info"].get("ticket_id", "?"), "task": state["info"].get("task", ""), "score": reward, "time": time.strftime("%H:%M:%S")})

    if reward >= 0.6:
        notify(f"✅ Passed! Score: {reward:.2f}", "success")
    else:
        notify(f"❌ Failed. Score: {reward:.2f}", "error")

    return gr.update(visible=True, value=card("📋 Results", results_html, status_color)), action_md


def auto_inspect():
    if state["observation"] is None:
        return gr.update(), gr.update(), gr.update(), gr.update(), gr.update()
    obs = state["observation"]
    action = _build_heuristic_action(obs.ticket, obs.worker_response)
    return gr.update(value=action.flagged), gr.update(value=", ".join(action.flagged_fields)), gr.update(value=json.dumps(action.issues, indent=2)), gr.update(value=json.dumps(action.suggested_corrections, indent=2)), gr.update(value=action.confidence)


def auto_inspect_dataset():
    if "_amazon_ticket" not in state:
        return gr.update(), gr.update(), gr.update(), gr.update(), gr.update()
    ticket = state["_amazon_ticket"]
    worker = state["_amazon_worker"]
    action = _build_heuristic_action(ticket, worker)
    return gr.update(value=action.flagged), gr.update(value=", ".join(action.flagged_fields)), gr.update(value=json.dumps(action.issues, indent=2)), gr.update(value=json.dumps(action.suggested_corrections, indent=2)), gr.update(value=action.confidence)


def run_benchmark(tickets_per_task, progress=gr.Progress()):
    bench_env = FleetAIEnv()
    data = {}
    for idx, task in enumerate(["inspection_easy", "inspection_hard", "inspection_adversarial"]):
        progress((idx * tickets_per_task) / (3 * tickets_per_task), f"Running {TASK_LABELS[task]}...")
        scores = []
        for seed in range(tickets_per_task):
            r = bench_env.reset(ResetRequest(task=task, seed=seed))
            action = _build_heuristic_action(r.observation.ticket, r.observation.worker_response)
            scores.append(bench_env.step(action).reward)
        data[task] = scores
    progress(1.0, "Done!")

    rows = ""
    all_scores = []
    for task in ["inspection_easy", "inspection_hard", "inspection_adversarial"]:
        scores = data[task]
        mean_val = sum(scores) / len(scores) if scores else 0
        all_scores.extend(scores)
        color = "#22c55e" if mean_val >= 0.7 else "#f59e0b" if mean_val >= 0.4 else "#ef4444"
        rows += f'<tr><td style="padding:6px 8px; color:#e2e8f0; border-bottom:1px solid #2a2a4a;"><span style="color:{TASK_COLORS[task]};">●</span> {TASK_LABELS[task]}</td><td style="padding:6px 8px; border-bottom:1px solid #2a2a4a;"><div style="display:flex; align-items:center; gap:6px;"><div style="flex:1;">{progress_bar(mean_val, color)}</div><span style="color:#e2e8f0; font-weight:600; font-size:12px;">{mean_val:.3f}</span></div></td><td style="padding:6px 8px; color:#8888aa; border-bottom:1px solid #2a2a4a; font-size:12px;">{min(scores):.3f}-{max(scores):.3f}</td></tr>'

    overall = sum(all_scores) / len(all_scores) if all_scores else 0
    oc = "#22c55e" if overall >= 0.7 else "#f59e0b" if overall >= 0.4 else "#ef4444"
    notify(f"Benchmark: {overall:.3f} avg ({len(all_scores)} tickets)", "success")
    return (
        f'<div style="display:flex; justify-content:space-between; align-items:center; margin-bottom:12px;">'
        f'<h4 style="margin:0; color:#e2e8f0;">Overall: <span style="color:{oc}; font-weight:700; font-size:20px;">{overall:.3f}</span></h4>'
        f'<span style="color:#8888aa; font-size:12px;">{len(all_scores)} tickets</span></div>'
        f'<table style="width:100%; border-collapse:collapse; font-size:13px;"><tr style="color:#8888aa; border-bottom:1px solid #4a4a6a;"><th style="text-align:left; padding:6px 8px;">Difficulty</th><th style="text-align:left; padding:6px 8px;">Score</th><th style="text-align:left; padding:6px 8px;">Range</th></tr>{rows}</table>'
    )


def refresh_stats():
    total = state["total_inspected"]
    passed = state["total_passed"]
    failed = state["total_failed"]
    avg = state["score_sum"] / total if total > 0 else 0
    elapsed = int(time.time() - state["session_start"])
    pr = (passed / total * 100) if total > 0 else 0

    stats = (
        f'<div style="display:grid; grid-template-columns:repeat(auto-fit, minmax(100px, 1fr)); gap:8px; margin-bottom:12px;">'
        f'{metric_box(total, "Inspected", "#3b82f6")}'
        f'{metric_box(passed, "Passed", "#22c55e")}'
        f'{metric_box(failed, "Failed", "#ef4444")}'
        f'{metric_box(f"{avg:.2f}", "Avg Score", "#f59e0b")}'
        f'{metric_box(f"{elapsed//60:02d}:{elapsed%60:02d}", "Session", "#8b5cf6")}'
        f'</div>'
        f'<div style="padding:8px; background:#12122a; border-radius:8px;"><div style="display:flex; justify-content:space-between; margin-bottom:4px;"><span style="color:#8888aa; font-size:11px;">Pass Rate</span><span style="color:#e2e8f0; font-weight:600; font-size:12px;">{pr:.0f}%</span></div>{progress_bar(pr/100, "#22c55e")}</div>'
    )

    hist_rows = ""
    for h in reversed(state["history"][-15:]):
        t = h.get("task", "")
        sc = "#22c55e" if h["score"] >= 0.6 else "#f59e0b" if h["score"] >= 0.4 else "#ef4444"
        hist_rows += f'<tr><td style="padding:3px 6px; color:#8888aa; font-size:11px;">{h["time"]}</td><td style="padding:3px 6px; color:#e2e8f0; font-size:11px;">{h["ticket_id"]}</td><td style="padding:3px 6px; color:#8888aa; font-size:11px;">{TASK_LABELS.get(t, t)}</td><td style="padding:3px 6px; color:{sc}; font-weight:600; font-size:11px;">{h["score"]:.3f}</td></tr>'
    if not hist_rows:
        hist_rows = '<tr><td colspan="4" style="padding:8px; color:#8888aa; text-align:center; font-size:12px;">No inspections yet</td></tr>'

    history = (
        f'<table style="width:100%; border-collapse:collapse;"><tr style="color:#8888aa; border-bottom:1px solid #4a4a6a;">'
        f'<th style="text-align:left; padding:3px 6px; font-size:11px;">Time</th><th style="text-align:left; padding:3px 6px; font-size:11px;">Ticket</th>'
        f'<th style="text-align:left; padding:3px 6px; font-size:11px;">Difficulty</th><th style="text-align:left; padding:3px 6px; font-size:11px;">Score</th></tr>'
        f'{hist_rows}</table>'
    )

    notifs = ""
    for n in state["notifications"][:10]:
        icon = {"success": "✅", "error": "❌", "info": "ℹ️"}.get(n["level"], "ℹ️")
        nc = {"success": "#22c55e", "error": "#ef4444", "info": "#3b82f6"}.get(n["level"], "#8888aa")
        notifs += f'<div style="padding:4px 6px; border-left:3px solid {nc}; margin-bottom:3px; background:#12122a; border-radius:3px; font-size:12px;"><span style="color:#8888aa;">{n["time"]}</span> {icon} <span style="color:#e2e8f0;">{n["message"]}</span></div>'
    if not notifs:
        notifs = '<span style="color:#8888aa; font-size:12px;">No notifications</span>'

    return stats, history, notifs


def _normalize_headers(headers):
    h_map = {}
    h_lower = [h.lower().strip() if h else "" for h in headers]
    for i, hl in enumerate(h_lower):
        if hl == "ticket_id":
            h_map["ticket_id"] = i
        elif hl == "order_id":
            h_map["order_id"] = i
        elif hl == "customer_id":
            h_map["customer_id"] = i
        elif hl == "marketplace":
            h_map["marketplace"] = i
        elif hl == "issue_type":
            h_map["issue_type"] = i
        elif hl == "sub_issue":
            h_map["sub_issue"] = i
        elif hl == "description":
            h_map["description"] = i
        elif hl == "priority":
            h_map["priority"] = i
        elif hl == "status":
            h_map["status"] = i
        elif hl == "created_date":
            h_map["created_date"] = i
        elif hl == "resolved_date":
            h_map["resolved_date"] = i
        elif hl == "assigned_agent":
            h_map["assigned_agent"] = i
        elif hl == "channel":
            h_map["channel"] = i
        elif hl == "amount_usd":
            h_map["amount_usd"] = i
        elif hl == "sla_hours":
            h_map["sla_hours"] = i
        elif hl == "customer_rating":
            h_map["customer_rating"] = i
        elif hl == "duplicate_flag":
            h_map["duplicate_flag"] = i
        elif hl == "notes":
            h_map["notes"] = i
    return h_map


def upload_dataset(file):
    if file is None:
        return gr.update(value="❌ No file selected. Please upload an Excel file."), "", ""

    import openpyxl
    try:
        wb = openpyxl.load_workbook(file)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        rows = list(ws.iter_rows(min_row=2, values_only=True))
    except Exception as e:
        return gr.update(value=f"❌ Failed to read file: {e}"), "", ""

    h_map = _normalize_headers(headers)
    if "ticket_id" not in h_map:
        return gr.update(value=f'❌ No ticket ID column found. Columns: {", ".join(str(h) for h in headers[:10])}'), "", ""

    total = len(rows)
    ticket_types = {}
    for r in rows:
        idx = h_map.get("issue_type")
        tt = str(r[idx] if idx is not None and idx < len(r) else "Unknown")
        ticket_types[tt] = ticket_types.get(tt, 0) + 1

    type_rows = ""
    for it, count in sorted(ticket_types.items(), key=lambda x: -x[1]):
        type_rows += f'<tr><td style="padding:4px 8px; color:#e2e8f0; font-size:12px;">{it}</td><td style="padding:4px 8px; color:#8888aa; font-size:12px;">{count}</td></tr>'

    preview_rows = ""
    for i, r in enumerate(rows[:5]):
        ti = h_map["ticket_id"]
        tid = str(r[ti] if ti < len(r) else f"ROW-{i+1}")
        it_i = h_map.get("issue_type")
        tt = str(r[it_i] if it_i is not None and it_i < len(r) else "N/A")
        desc_i = h_map.get("description")
        desc_short = str(r[desc_i] if desc_i is not None and desc_i < len(r) else "")[:40]
        preview_rows += f'<tr><td style="padding:3px 6px; color:#e2e8f0; font-size:11px;">{tid}</td><td style="padding:3px 6px; color:#a78bfa; font-size:11px;">{tt}</td><td style="padding:3px 6px; color:#8888aa; font-size:11px;">{desc_short}</td></tr>'
    more = f'<tr><td colspan="3" style="padding:4px; color:#8888aa; text-align:center; font-size:11px;">... and {total - 5} more</td></tr>' if total > 5 else ""

    summary = (
        f'<div style="padding:10px; background:#12122a; border-radius:8px; margin-bottom:8px;">'
        f'<div style="font-size:14px; color:#e2e8f0; margin-bottom:4px;">{file.name}</div>'
        f'<div style="color:#8888aa; font-size:12px;">{total} tickets | {len(ticket_types)} types | {len(headers)} columns</div></div>'
        f'<div style="padding:8px; background:#12122a; border-radius:8px; margin-bottom:8px;"><table style="width:100%; border-collapse:collapse; font-size:12px;">'
        f'<tr style="color:#8888aa; border-bottom:1px solid #4a4a6a;"><th style="text-align:left; padding:4px 8px;">Type</th><th style="text-align:left; padding:4px 8px;">Count</th></tr>'
        f'{type_rows}</table></div>'
        f'<div style="padding:8px; background:#12122a; border-radius:8px;"><div style="color:#8888aa; font-size:11px; margin-bottom:4px;">PREVIEW (first 5)</div>'
        f'<table style="width:100%; border-collapse:collapse; font-size:11px;">'
        f'<tr style="color:#8888aa; border-bottom:1px solid #4a4a6a;"><th style="text-align:left; padding:3px 6px;">ID</th><th style="text-align:left; padding:3px 6px;">Type</th><th style="text-align:left; padding:3px 6px;">Description</th></tr>'
        f'{preview_rows}{more}</table></div>'
    )

    state["dataset_file"] = file.name
    state["amazon_rows"] = rows
    state["amazon_headers"] = headers
    state["amazon_hmap"] = h_map
    state["amazon_index"] = 0
    notify(f"Loaded {file.name}: {total} tickets", "success")

    return summary, file.name, total


def load_dataset_ticket(index):
    if "amazon_rows" not in state or not state["amazon_rows"]:
        return gr.update(value="No dataset loaded. Upload a file first."), "", "", gr.update(), gr.update()

    rows = state["amazon_rows"]
    headers = state["amazon_headers"]
    h_map = state.get("amazon_hmap", {})

    if index < 0 or index >= len(rows):
        return gr.update(value=f"Index {index} out of range (0-{len(rows)-1})"), "", "", gr.update(), gr.update()

    state["amazon_index"] = index
    row = rows[index]

    def g(key, default=""):
        idx = h_map.get(key)
        if idx is not None and idx < len(row):
            return str(row[idx]) if row[idx] is not None else default
        return default

    tid = g("ticket_id", f"ROW-{index+1}")
    issue_type = g("issue_type", "General")
    sub_issue = g("sub_issue", "")
    desc = g("description", "")
    subject = f"[{issue_type}] {sub_issue}"
    priority_raw = g("priority", "Medium").strip()
    customer_id = g("customer_id", "C-000")
    notes = g("notes", "")
    status = g("status", "")
    order_id = g("order_id", "")
    marketplace = g("marketplace", "")
    channel = g("channel", "")
    assigned_agent = g("assigned_agent", "")
    customer_rating = g("customer_rating", "")
    sla_hours = g("sla_hours", "")

    classification = ISSUE_TYPE_TO_CLASSIFICATION.get(issue_type, "general")
    priority = PRIORITY_MAP.get(priority_raw, "medium")

    negative_words = ["angry", "frustrated", "terrible", "worst", "refund", "complaint", "overcharged", "not received", "stolen", "broken", "damage", "poor", "unacceptable", "disappointed", "worst", "late", "missing"]
    sentiment = "negative" if any(w in f"{desc} {notes} {subject}".lower() for w in negative_words) else "neutral"

    from environment.error_injector import ErrorInjector
    from environment.tickets import TicketRecord, TicketGroundTruth
    from environment.graders import grade_inspector

    gt = TicketGroundTruth(category=classification, priority=priority, department="general_support")
    ticket = TicketRecord(id=tid, subject=subject, body=desc, customer_id=customer_id, customer_name=customer_id, customer_tier=TIER_MAP.get(notes.lower(), "silver"), customer_tenure_days=180, sentiment=sentiment, previous_ticket_count=2, ground_truth=gt)

    rng = random.Random(index + 42)
    injector = ErrorInjector(rng=rng)
    worker, errors = injector.generate_worker_response(ticket)

    state["_amazon_ticket"] = ticket
    state["_amazon_worker"] = worker
    state["_amazon_errors"] = errors

    sent_emoji = {"negative": "😡", "very_negative": "🤬", "neutral": "😐", "positive": "😊"}.get(sentiment, "😐")
    fail_html = '<span style="color:#ef4444;">None</span>'
    priority_color = "#ef4444" if worker.priority == "urgent" else "#f59e0b" if worker.priority == "high" else "#22c55e"

    ticket_html = card("Ticket", (
        f'<div style="display:grid; grid-template-columns:1fr 1fr 1fr; gap:8px; margin-bottom:10px;">'
        f'<div style="padding:8px; background:#12122a; border-radius:8px;"><span style="color:#8888aa; font-size:11px;">Ticket ID</span><br><span style="color:#e2e8f0; font-weight:600;">{tid}</span></div>'
        f'<div style="padding:8px; background:#12122a; border-radius:8px;"><span style="color:#8888aa; font-size:11px;">Order ID</span><br><span style="color:#e2e8f0; font-weight:600;">{order_id}</span></div>'
        f'<div style="padding:8px; background:#12122a; border-radius:8px;"><span style="color:#8888aa; font-size:11px;">Customer</span><br><span style="color:#e2e8f0; font-weight:600;">{customer_id}</span></div>'
        f'<div style="padding:8px; background:#12122a; border-radius:8px;"><span style="color:#8888aa; font-size:11px;">Type</span><br><span style="color:#a78bfa; font-weight:600;">{issue_type}</span></div>'
        f'<div style="padding:8px; background:#12122a; border-radius:8px;"><span style="color:#8888aa; font-size:11px;">Priority</span><br><span style="color:{priority_color}; font-weight:600;">{priority_raw}</span></div>'
        f'<div style="padding:8px; background:#12122a; border-radius:8px;"><span style="color:#8888aa; font-size:11px;">Status</span><br><span style="color:#e2e8f0; font-weight:600;">{status}</span></div></div>'
        f'<div style="padding:10px; background:#12122a; border-radius:8px; margin-bottom:6px;"><span style="color:#8888aa; font-size:11px;">Description</span><br><span style="color:#e2e8f0;">{desc[:500]}</span></div>'
        f'<div style="display:grid; grid-template-columns:1fr 1fr 1fr 1fr; gap:8px;">'
        f'<div style="padding:6px; background:#12122a; border-radius:8px;"><span style="color:#8888aa; font-size:10px;">Marketplace</span><br><span style="color:#e2e8f0; font-size:12px;">{marketplace}</span></div>'
        f'<div style="padding:6px; background:#12122a; border-radius:8px;"><span style="color:#8888aa; font-size:10px;">Channel</span><br><span style="color:#e2e8f0; font-size:12px;">{channel}</span></div>'
        f'<div style="padding:6px; background:#12122a; border-radius:8px;"><span style="color:#8888aa; font-size:10px;">Agent</span><br><span style="color:#e2e8f0; font-size:12px;">{assigned_agent}</span></div>'
        f'<div style="padding:6px; background:#12122a; border-radius:8px;"><span style="color:#8888aa; font-size:10px;">Sentiment</span><br><span style="font-size:14px;">{sent_emoji} {sentiment.title()}</span></div></div>'
    ), "#a78bfa")

    worker_html = card("Worker Response", (
        f'<div style="display:grid; grid-template-columns:1fr 1fr 1fr; gap:8px; margin-bottom:10px;">'
        f'<div style="padding:8px; background:#12122a; border-radius:8px;"><span style="color:#8888aa; font-size:11px;">Classification</span><br><span style="color:#a78bfa; font-weight:600;">{worker.classification or "N/A"}</span></div>'
        f'<div style="padding:8px; background:#12122a; border-radius:8px;"><span style="color:#8888aa; font-size:11px;">Priority</span><br><span style="color:{priority_color}; font-weight:600;">{worker.priority or "N/A"}</span></div>'
        f'<div style="padding:8px; background:#12122a; border-radius:8px;"><span style="color:#8888aa; font-size:11px;">Department</span><br><span style="color:#38bdf8; font-weight:600;">{worker.department or "N/A"}</span></div></div>'
        f'<div style="padding:10px; background:#12122a; border-radius:8px; margin-bottom:6px;"><span style="color:#8888aa; font-size:11px;">Response</span><br><span style="color:#e2e8f0;">{(worker.response or "N/A")[:400]}</span></div>'
        f'<div style="padding:8px; background:#12122a; border-radius:8px;"><span style="color:#8888aa; font-size:11px;">Resolution Actions</span><br><span style="color:#e2e8f0;">{", ".join(worker.resolution_actions) if worker.resolution_actions else fail_html}</span></div>'
    ), "#8b5cf6")

    error_fields = [e.field for e in errors]
    errors_str = ", ".join(error_fields) if error_fields else "none"

    info_text = f"Ticket {tid} ({issue_type}) loaded.\nInjected errors: {errors_str}"

    notify(f"Ticket loaded: {tid}", "info")
    return gr.update(), ticket_html, worker_html, info_text, gr.update(minimum=0, maximum=len(rows)-1, value=index)


def submit_dataset_inspection(flagged, flagged_fields, issues_json, corrections_json, confidence):
    if "_amazon_ticket" not in state:
        return gr.update(visible=True, value="❌ Load a ticket first!"), ""

    from environment.graders import grade_inspector

    ticket = state["_amazon_ticket"]
    errors = state["_amazon_errors"]

    parsed_fields = [f.strip() for f in flagged_fields.split(",") if f.strip()] if flagged_fields.strip() else []
    if flagged and not parsed_fields:
        return gr.update(visible=True, value="⚠️ Flagged=True but no fields listed."), ""

    parsed_issues = []
    if issues_json.strip():
        try:
            parsed_issues = json.loads(issues_json.strip())
        except json.JSONDecodeError as e:
            return gr.update(visible=True, value=f"❌ Invalid JSON: {e}"), ""

    parsed_corrections = {}
    if corrections_json.strip():
        try:
            parsed_corrections = json.loads(corrections_json.strip())
        except json.JSONDecodeError as e:
            return gr.update(visible=True, value=f"❌ Invalid JSON: {e}"), ""

    action = InspectorAction(flagged=flagged, flagged_fields=parsed_fields, issues=parsed_issues, suggested_corrections=parsed_corrections, confidence=confidence)
    grader_result = grade_inspector(errors, action, ticket)
    reward = grader_result.overall_score

    if reward >= 0.8:
        status_label, status_color = "🏆 EXCELLENT", "#22c55e"
    elif reward >= 0.6:
        status_label, status_color = "✅ GOOD", "#22c55e"
    elif reward >= 0.4:
        status_label, status_color = "⚠️ OK", "#f59e0b"
    elif reward >= 0.2:
        status_label, status_color = "❌ NEEDS WORK", "#f59e0b"
    else:
        status_label, status_color = "💀 FAIL", "#ef4444"

    comp_html = ""
    for cs in grader_result.component_scores:
        bar_c = "#22c55e" if cs.score >= 0.7 else "#f59e0b" if cs.score >= 0.4 else "#ef4444"
        comp_html += f'<div style="margin-bottom:6px;"><div style="display:flex; justify-content:space-between; margin-bottom:2px;"><span style="color:#e2e8f0; font-size:12px;">{cs.name.replace("_", " ").title()}</span><span style="color:#8888aa; font-size:11px;">{cs.score:.2f}</span></div>{progress_bar(cs.score, bar_c)}</div>'

    err_rows = ""
    for e in errors:
        gt_val = e.original_value if hasattr(e, "original_value") else ""
        corrupted_val = e.corrupted_value if hasattr(e, "corrupted_value") else ""
        err_rows += f'<tr><td style="padding:4px 8px; color:#e2e8f0; border-bottom:1px solid #2a2a4a;">{e.field}</td><td style="padding:4px 8px; color:#8888aa; border-bottom:1px solid #2a2a4a;">{e.error_type.value if hasattr(e.error_type, "value") else str(e.error_type)}</td></tr>'
    if not err_rows:
        err_rows = '<tr><td colspan="3" style="padding:4px; color:#22c55e; text-align:center;">✅ No errors (clean response)</td></tr>'

    results_html = (
        f'<div style="text-align:center; padding:16px; background:#12122a; border-radius:10px; margin-bottom:12px;">'
        f'<div style="font-size:32px; font-weight:700; color:{status_color};">{reward:.2f}</div>'
        f'<div style="color:#8888aa; font-size:13px;">{status_label}</div></div>'
        f'{comp_html}'
        f'<div style="margin-top:8px;"><span style="color:#8888aa; font-size:11px;">INJECTED ERRORS</span>'
        f'<table style="width:100%; border-collapse:collapse; font-size:12px; margin-top:4px;"><tr style="color:#8888aa;"><th style="text-align:left; padding:4px 8px;">Field</th><th style="text-align:left; padding:4px 8px;">Type</th></tr>{err_rows}</table></div>'
    )

    action_md = f"```json\n{json.dumps(action.model_dump(), indent=2)}\n```"

    state["total_inspected"] += 1
    state["score_sum"] += reward
    if reward >= 0.4:
        state["total_passed"] += 1
    else:
        state["total_failed"] += 1
    state["history"].append({"ticket_id": state.get("dataset_file", "?"), "task": "dataset", "score": reward, "time": time.strftime("%H:%M:%S")})

    if reward >= 0.6:
        notify(f"✅ {tid} passed! Score: {reward:.2f}", "success")
    else:
        notify(f"❌ {tid} failed. Score: {reward:.2f}", "error")

    return gr.update(visible=True, value=card("Results", results_html, status_color)), action_md


CSS = """
@media (max-width: 768px) {
    .gr-form { gap: 8px !important; }
    .gr-row { flex-direction: column !important; }
    .gr-column { width: 100% !important; }
    .gr-container { padding: 10px !important; }
}
.gradio-container {
    max-width: 960px !important;
    margin: 0 auto !important;
}
"""


with gr.Blocks(title="FleetAI Dashboard") as demo:

    gr.HTML('<div style="text-align:center; padding:16px; background:linear-gradient(135deg, #1a1a2e, #16213e); border-radius:12px; margin-bottom:16px;"><h1 style="margin:0; color:#f1f5f9; font-size:24px;">🛡️ FleetAI - Scalable Oversight</h1><p style="margin:4px 0 0 0; color:#8888aa; font-size:14px;">AI Inspector Dashboard for Customer Support Quality Assurance</p></div>')

    with gr.Tabs():
        with gr.Tab("Dashboard"):
            stats_display = gr.HTML(value=refresh_stats()[0])
            gr.Markdown("### Recent History")
            history_display = gr.HTML(value=refresh_stats()[1])
            gr.Markdown("### Notifications")
            notifications_display = gr.HTML(value=refresh_stats()[2])

        with gr.Tab("Inspect Tickets"):
            gr.Markdown("*Configure, start, inspect, and get scored in one flow.*")

            with gr.Row():
                with gr.Column(scale=1):
                    task_desc_md = gr.Markdown(TASK_DESCRIPTIONS["inspection_easy"])
                    with gr.Row():
                        task_dropdown = gr.Dropdown(choices=list(TASK_LABELS.keys()), value="inspection_easy", label="Difficulty")
                        seed_input = gr.Number(value=42, label="Seed", precision=0)
                    start_btn = gr.Button("🚀 Start Episode", variant="primary")
                    with gr.Row():
                        auto_btn = gr.Button("Auto-Inspect")
                        submit_btn = gr.Button("Submit Inspection", variant="primary")
                    flagged_check = gr.Checkbox(label="Flagged", value=False)
                    flagged_fields_input = gr.Textbox(label="Flagged Fields", placeholder="priority, response", info="Comma-separated")
                    issues_input = gr.Textbox(label="Issues JSON", placeholder='[{"field":"priority","reason":"Should be urgent"}]', lines=3)
                    corrections_input = gr.Textbox(label="Corrections JSON", placeholder='{"priority":"urgent"}', lines=2)
                    confidence_slider = gr.Slider(0, 1, value=0.5, step=0.05, label="Confidence")

                with gr.Column(scale=2):
                    ticket_display = gr.HTML('<div style="color:#8888aa; text-align:center; padding:40px; background:#1a1a2e; border-radius:12px;">Click "Start Episode" to load a ticket</div>')
                    worker_display = gr.HTML("")
                    instructions_display = gr.Textbox(label="Task Instructions", lines=3, interactive=False, visible=False)

            with gr.Row():
                status_display = gr.HTML(visible=True)
                results_display = gr.HTML(visible=False)
                action_display = gr.Markdown("")

        with gr.Tab("Dataset"):
            gr.Markdown("*Upload your Excel file (.xlsx) to inspect tickets through the FleetAI grader.*")

            with gr.Row():
                upload = gr.File(label="Upload Excel File", file_types=[".xlsx"], type="filepath")
                upload_btn = gr.Button("Load Dataset", variant="primary")
            upload_summary = gr.HTML('<div style="color:#8888aa; font-size:13px; text-align:center; padding:20px; background:#1a1a2e; border-radius:12px;">Upload an Excel file to begin</div>')

            with gr.Row():
                with gr.Column(scale=1):
                    ticket_slider = gr.Slider(0, 0, value=0, step=1, label="Ticket Index", visible=False)
                    load_btn = gr.Button("Load This Ticket", visible=False)
                    upload_info = gr.Textbox(label="Info", lines=2, interactive=False, visible=False)
                with gr.Column(scale=2):
                    ticket_display2 = gr.HTML()
                    worker_display2 = gr.HTML()

            ds_flagged_check = gr.Checkbox(label="Flagged", value=False)
            ds_flagged_fields = gr.Textbox(label="Flagged Fields", placeholder="priority, response", info="Comma-separated")
            ds_issues = gr.Textbox(label="Issues JSON", placeholder='[{"field":"priority","reason":"Should be urgent"}]', lines=3)
            ds_corrections = gr.Textbox(label="Corrections JSON", placeholder='{"priority":"urgent"}', lines=2)
            ds_confidence = gr.Slider(0, 1, value=0.5, step=0.05, label="Confidence")

            with gr.Row():
                ds_auto_btn = gr.Button("Auto-Inspect")
                ds_submit_btn = gr.Button("Submit Inspection", variant="primary")

            ds_status_display = gr.HTML(visible=True)
            ds_results_display = gr.HTML(visible=False)
            ds_action_display = gr.Markdown("")

        with gr.Tab("Benchmark"):
            gr.Markdown("*Evaluate the heuristic inspector across all difficulty levels.*")
            with gr.Row():
                benchmark_tickets = gr.Slider(3, 30, value=10, step=1, label="Tickets per task")
                benchmark_btn = gr.Button("Run Benchmark", variant="primary")
            benchmark_output = gr.HTML("")

        with gr.Tab("About"):
            gr.Markdown("""
            ## FleetAI - Scalable Oversight

            ### The Problem
            As companies deploy LLM-based agents for customer support, there is no standardized way to **train and evaluate oversight agents** that audit their outputs. **Who watches the watchers?**

            ### The Solution
            FleetAI creates a **two-agent ecosystem**: a Worker agent handles support tickets, and an Inspector agent reviews the Worker's output for errors, policy violations, and quality issues.

            ### Key Features
            1. **Error Injection** - Realistic Worker mistakes at varying subtlety levels (obvious, subtle, multi-error, clean traps)
            2. **5-Component Grading** - Error Detection (35%), Precision (25%), Issue Specificity (15%), Correction Quality (15%), Calibration (10%)
            3. **Anti-Hack Checks** - Penalizes flag-everything, max-confidence, and copy-paste strategies
            4. **3 Difficulty Levels** - Easy, Hard, Adversarial
            5. **GRPO Training** - SFT warmstart → Reinforcement Learning with progressive difficulty

            ### How to Use
            1. **Inspect Tickets** - Start episodes, review worker responses, submit inspections
            2. **Dataset** - Upload your Excel file, browse tickets, inspect each one
            3. **Benchmark** - Run batch evaluation to see overall performance
            """)

    task_dropdown.change(fn=lambda t: gr.update(value=TASK_DESCRIPTIONS.get(t, "")), inputs=[task_dropdown], outputs=[task_desc_md])
    start_btn.click(fn=start_episode, inputs=[task_dropdown, seed_input], outputs=[ticket_display, worker_display, instructions_display, results_display, status_display, action_display])
    auto_btn.click(fn=auto_inspect, outputs=[flagged_check, flagged_fields_input, issues_input, corrections_input, confidence_slider])
    submit_btn.click(fn=submit_inspection, inputs=[flagged_check, flagged_fields_input, issues_input, corrections_input, confidence_slider], outputs=[results_display, action_display])
    upload_btn.click(fn=upload_dataset, inputs=[upload], outputs=[upload_summary, upload, ticket_slider])
    upload_summary.change(fn=lambda: (gr.update(visible=True), gr.update(visible=True), gr.update(visible=True)), inputs=[], outputs=[ticket_slider, load_btn, upload_info])
    load_btn.click(fn=load_dataset_ticket, inputs=[ticket_slider], outputs=[ds_status_display, ticket_display2, worker_display2, upload_info, ticket_slider])
    ds_auto_btn.click(fn=auto_inspect_dataset, outputs=[ds_flagged_check, ds_flagged_fields, ds_issues, ds_corrections, ds_confidence])
    ds_submit_btn.click(fn=submit_dataset_inspection, inputs=[ds_flagged_check, ds_flagged_fields, ds_issues, ds_corrections, ds_confidence], outputs=[ds_results_display, ds_action_display])
    benchmark_btn.click(fn=run_benchmark, inputs=[benchmark_tickets], outputs=[benchmark_output])


if __name__ == "__main__":
    demo.launch(server_name="0.0.0.0", server_port=7862, share=False, css=CSS)
